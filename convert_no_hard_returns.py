import os
import time
import re
import pdfplumber
from openpyxl import Workbook

# Number of reports extracted
n_reports = 0

# Start the timer
start_time = time.time()

# Set your directory containing the PDF files
#pdf_folder = "pdfs"
pdf_folder = "test_pdf"
output_excel = "pdf_texts.xlsx"

# Create a new Excel workbook
wb = Workbook()
# Remove the default sheet
wb.remove(wb.active)

# Loop through each PDF file in the directory
for filename in os.listdir(pdf_folder):
    if filename.lower().endswith(".pdf"):
        pdf_path = os.path.join(pdf_folder, filename)
        sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet name max length is 31

        # Extract text from PDF
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n\n"  # ensure page breaks contribute to paragraph separation

        # Add a new sheet for each PDF
        ws = wb.create_sheet(title=sheet_name)

        # Split the text into paragraphs and write to Excel
        paragraphs = re.split(r'\n\s*\n', text)
        for i, para in enumerate(paragraphs, start=1):
            clean_para = para.strip()
            if clean_para:
                # Replace line breaks within paragraphs with space
                clean_para = clean_para.replace('\n', ' ')
                # Optional: normalize whitespace
                clean_para = re.sub(r'\s+', ' ', clean_para).strip()
                # Write to Excel
                ws.cell(row=i, column=1, value=clean_para)

        # Increment the report count
        n_reports += 1
        print(f"{n_reports} PDF(s) extracted to Excel: {filename}")

# Save the workbook
wb.save(output_excel)

# End the timer
end_time = time.time()
elapsed_time = (end_time - start_time) / 60

print(f"\nText from PDFs saved to '{output_excel}'")
print(f"Total PDFs processed: {n_reports}")
print(f"Process completed in {elapsed_time:.2f} minutes.")
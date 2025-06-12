import os
import time
import re
import pdfplumber
from openpyxl import Workbook


def fix_split_number_lines(text):
    """
    Fix lines that contain only digits followed by a numbered paragraph (e.g., '6\\n7.') into one line (e.g., '67.').
    """
    lines = text.splitlines()
    cleaned_lines = []
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if line.isdigit() and i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            match = re.match(r'^(\d+)\.(.*)', next_line)
            if match:
                merged_line = line + match.group(1) + "." + match.group(2)
                cleaned_lines.append(merged_line.strip())
                i += 2
                continue
        cleaned_lines.append(line)
        i += 1

    return '\n'.join(cleaned_lines)


def clean_text(raw_text):
    """
    Remove control characters and normalize whitespace.
    """
    raw_text = re.sub(r'[\x00-\x1F\x7F]', ' ', raw_text)
    return re.sub(r'\s+', ' ', raw_text).strip()


def split_paragraphs(text):
    """
    Split text into numbered paragraphs like '21. ...', '22. ...'
    """
    return re.split(r'(?=\b\d{1,3}\.\s)', text)


def convert_pdfs_to_excel_paragraphs(pdf_folder, output_excel):
    """
    Converts cleaned and numbered-paragraph PDFs into an Excel file.
    Each paragraph is placed into its own cell (row) in the first column.
    One sheet per PDF.
    """
    n_reports = 0
    start_time = time.time()

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for filename in os.listdir(pdf_folder):
        if filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, filename)
            sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet name limit

            # Extract raw text from PDF
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"

            # Step 1: Fix broken paragraph numbers
            text = fix_split_number_lines(text)

            # Step 2: Fix patterns like 1\n2. or 1\n1\n9.
            text = re.sub(r'(?:\n)?(?<!\d)(\d)\n(\d)\n(\d)\.', r'\1\2\3.', text)
            text = re.sub(r'(?:\n)?(?<!\d)(\d)\n(\d)\.', r'\1\2.', text)

            # Step 3: Merge non-paragraph-breaking line breaks into spaces
            text = re.sub(r'(?<!\n)\n(?!\n)', ' ', text)

            # Step 4: Clean and normalize the text
            cleaned_text = clean_text(text)

            # Step 5: Split into paragraphs
            paragraphs = split_paragraphs(cleaned_text)

            # Step 6: Write to Excel
            ws = wb.create_sheet(title=sheet_name)
            row = 1
            for para in paragraphs:
                clean_para = para.strip()
                if clean_para:
                    ws.cell(row=row, column=1, value=clean_para)
                    row += 1

            n_reports += 1
            print(f"{n_reports} PDF(s) converted to Excel: {filename}")

    # Save Excel file
    wb.save(output_excel)
    elapsed_time = (time.time() - start_time) / 60

    print(f"\nText from PDFs saved to '{output_excel}'")
    print(f"Total PDFs processed: {n_reports}")
    print(f"Process completed in {elapsed_time:.2f} minutes.")


# Example usage:
convert_pdfs_to_excel_paragraphs(pdf_folder="pdfs", output_excel="numbered_paragraphs.xlsx")
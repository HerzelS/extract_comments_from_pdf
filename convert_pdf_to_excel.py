import os
import time
import pdfplumber
from openpyxl import Workbook


# Number of reports extrated
n_reports = 0

# Start the timer
start_time = time.time()

# Set your directory containing the PDF files
pdf_folder = "pdfs"
output_excel = "pdf_texts.xlsx"

# Create a new Excel workbook
wb = Workbook()

# Remove the default sheet
wb.remove(wb.active)

# Loop through each PDF file in the directory
for filename in os.listdir(pdf_folder):
    if filename.lower().endswith(".pdf"):
        pdf_path = os.path.join(pdf_folder, filename)
        sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet name max length is 31

        # Extract text from PDF
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""

        # Add a new sheet for each PDF
        ws = wb.create_sheet(title=sheet_name)

        # Split the text into lines and write to Excel
        for i, line in enumerate(text.splitlines(), start=1):
            ws.cell(row=i, column=1, value=line)
            
        # Increment the report count
        n_reports += 1
        print(f"A total of {n_reports} extracted from pdf to excel.")

# Save the workbook
wb.save(output_excel)

# End the timer
end_time = time.time()
elapsed_time = (end_time - start_time)/60

print(f"Text from PDFs saved to {output_excel}")
print(f"Process completed in {elapsed_time:.2f} minutes.")